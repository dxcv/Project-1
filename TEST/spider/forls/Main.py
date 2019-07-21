#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @CXM

#人行网站用js整体渲染反爬，第一次用selenium写爬虫

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import re
import openpyxl
import time
import random




#获取网页源代码 str

# print(source)
class Functions(object):

    def getcode(url):
        # 不显示webdriver弹出窗口
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('blink-settings=imagesEnabled=false')  # 不加载图片, 提升速度
        driver = webdriver.Chrome(chrome_options=chrome_options)  # 打开浏览器
        driver.implicitly_wait(10)
        driver.get(url)  # 用webdriver打开人行政策货币司
        time.sleep(2)
        nodemo = ['noscript']
        while nodemo != []:
            driver.refresh()
            source = driver.page_source  # 获取网页源代码 str
            pat_noscript = 'noscript'
            nodemo = re.compile(pat_noscript).findall(source)  # 防止js未加载完
            time.sleep(2)
        driver.close()  # 回收资源
        return source

    def getdetail(source):
        # 获取正文标题
        pat_title_article = '<h2 style="font-size: 16px;color: #333;">(.*?)</h2>'
        title_article = re.compile(pat_title_article).findall(source)

        ########正文内容#######
        # 查找日期
        pat_date = '<span id="shijian">(.*?)</span>'  # .replace(' ','')
        date = re.compile(pat_date).findall(source)[0][:10]

        # 查找通知正文部分
        pat_content = '<p>(.*?)</p>'
        content = re.compile(pat_content).findall(source)
        if content == [] or content[0] == '':
            pat_content = '<p align="left">(.*?)</p>'  # .replace(' ','')
            content = re.compile(pat_content).findall(source)
            detail = content[0]
        else:
            detail = content[0]

        # 查找页面中的表格
        pat_table = '<table align="center" border="1" cellpadding="0" cellspacing="0" style="border-bottom: windowtext 1px solid; border-left: windowtext 1px solid; width: 526px; border-top: windowtext 1px solid; border-right: windowtext 1px solid" width="526">([\s\S]*?)<tbody>([\s\S]*?)</tbody>([\s\S]*?)</table>'
        table = re.compile(pat_table).findall(source)
        table_alt_1 = []
        table_alt_2 = []
        table_alt_3 = []
        table_alt_4 = []
        if table == []:  # 添加特例
            pat_table_alt_1 = '<table align="center" border="1" cellpadding="0" cellspacing="0" style="border-right: windowtext 1px solid; border-top: windowtext 1px solid; border-left: windowtext 1px solid; width: 526px; border-bottom: windowtext 1px solid" width="526">([\s\S]*?)<tbody>([\s\S]*?)</tbody>([\s\S]*?)</table>'  # .replace(' ','')
            table_alt_1 = re.compile(pat_table_alt_1).findall(source)
            if table_alt_2 == []:
                pat_table_alt_2 = '<table align="center" border="1" cellpadding="0" cellspacing="0" style="border: 1px solid windowtext; width: 526px; margin:0 auto;" width="526">([\s\S]*?)<tbody>([\s\S]*?)</tbody>([\s\S]*?)</table>'  # .replace(' ','')
                table_alt_2 = re.compile(pat_table_alt_2).findall(source)
                if table_alt_2 == []:
                    pat_table_alt_3 = '<table align="center" border="1" cellpadding="0" cellspacing="0" style="border: 1px solid windowtext; width: 526px;" width="526">([\s\S]*?)<tbody>([\s\S]*?)</tbody>([\s\S]*?)</table>'
                    table_alt_3 = re.compile(pat_table_alt_3).findall(source)
                    if table_alt_3 == []:
                        pat_table_alt_4 = '<table align="center" border="1" cellpadding="0" cellspacing="0" style="border: 1px solid windowtext; width: 526px;" width="526">([\s\S]*?)<tbody>([\s\S]*?)</tbody>([\s\S]*?)</table>'
                        table_alt_4 = re.compile(pat_table_alt_4).findall(source)
                    else:
                        pass
                else:
                    pass
            else:
                pass
        else:
            pass

        detail_table_content = []
        if table == []:  # 通知中没有表格不做进一步的解析
            pass
        else:
            for t in table:
                # 列标题
                pat_title_table_column = '<span style="font-size: medium"><span style="color: #222222"><span style="font-family: 宋体">(.*?)</span>'  # .replace(' ','')
                title_table_column = re.compile(pat_title_table_column).findall(t[1])
                ###表格内容
                # 带汉字
                pat_table_content1 = '<span style="font-family: arial, sans-serif">(.*?)</span></span></span><span style="color: #222222"><span style="font-family: 宋体">(.*?)</span>'  # .replace(' ','')
                table_content1 = re.compile(pat_table_content1).findall(t[1])
                # 不带汉字
                pat_table_content2 = '<span style="font-family: arial, sans-serif"><span style="font-size: medium">(.*?)</span>'  # .replace(' ','')
                table_content2 = re.compile(pat_table_content2).findall(t[1])
                row = len(table_content2)
                if row == 0:
                    # 列标题
                    pat_title_table_column = '<span style="font-size: 12px"><span style="color: #222222"><span style="font-family: 宋体">(.*?)</span>'  # .replace(' ','')
                    title_table_column = re.compile(pat_title_table_column).findall(t[1])
                    ###表格内容
                    # 带汉字
                    pat_table_content1 = '<span style="font-size: 12px"><span style="color: #222222"><span style="font-family: arial, sans-serif">(.*?)</span><span style="font-family: 宋体">(.*?)</span></span></span>'  # .replace(' ','')
                    table_content1 = re.compile(pat_table_content1).findall(t[1])
                    # 不带汉字
                    pat_table_content2 = '<span style="font-size: 12px"><span style="color: #222222"><span style="font-family: arial, sans-serif">(.*?)</span></span></span>'  # .replace(' ','')
                    table_content2 = re.compile(pat_table_content2).findall(t[1])
                    row = int(len(table_content1) / 2)
                    if row == 0:
                        for t in table:
                            # 列标题
                            # 默认括号里的内容为group，要匹配括号里的文字要对括号转义
                            pat_title_table_column = '<span style="color: #222222"><span style="font-family: 宋体">(.*?)</span></span>'  # .replace(' ','')
                            title_table_column = re.compile(pat_title_table_column).findall(t[1])
                            ###表格内容
                            # 带汉字
                            pat_table_content1 = '<span style="color: #222222"><span style="font-family: arial, sans-serif">(.*?)</span></span><span style="color: #222222"><span style="font-family: 宋体">(.*?)</span>'  # .replace(' ','')
                            table_content1 = re.compile(pat_table_content1).findall(t[1])
                            # 不带汉字
                            pat_table_content2 = '<span style="color: #222222"><span style="font-family: arial, sans-serif">(.*?)</span>'  # .replace(' ','')
                            table_content2 = re.compile(pat_table_content2).findall(t[1])
                            row = int(len(table_content1) / 2)
                            for i in range(0, row):
                                if title_table_column[2] == '中标利率':
                                    detail_table_content.insert(len(detail_table_content),
                                                                '逆回购操作情况:' + title_table_column[0] + '为' + ''.join(
                                                                    table_content1[i * 2]) + "," + title_table_column[
                                                                    1] + '为' + ''.join(
                                                                    table_content1[i * 2 + 1]) + "," +
                                                                title_table_column[2] + '为' + table_content2[i])
                                elif title_table_column[2] == '操作利率':
                                    detail_table_content.insert(len(detail_table_content),
                                                                'MLF操作情况:' + title_table_column[0] + '为' + ''.join(
                                                                    table_content1[i * 2]) + "," + title_table_column[
                                                                    1] + '为' + ''.join(
                                                                    table_content1[i * 2 + 1]) + "," +
                                                                title_table_column[2] + '为' + table_content2[i])
                                else:
                                    detail_table_content.insert(len(detail_table_content), date + "当日情况存在异常")
                    else:
                        for i in range(0, row):
                            if title_table_column[2] == '中标利率':
                                detail_table_content.insert(len(detail_table_content),'逆回购操作情况:' + title_table_column[0] + '为' + ''.join(table_content1[i * 2]) + "," + title_table_column[1] + '为' + ''.join(table_content1[i * 2 + 1]) + "," +title_table_column[2] + '为' + table_content2[2])
                            elif title_table_column[2] == '操作利率':
                                detail_table_content.insert(len(detail_table_content),'MLF操作情况:' + title_table_column[0] + '为' + ''.join(table_content1[i * 2]) + "," + title_table_column[1] + '为' + ''.join(table_content1[i * 2 + 1]) + "," +title_table_column[2] + '为' + table_content2[2])
                            else:
                                detail_table_content.insert(len(detail_table_content), date + "当日情况存在异常")
                else:
                    for i in range(0, row):
                        if title_table_column[2] == '中标利率':
                            detail_table_content.insert(len(detail_table_content),'逆回购操作情况:' + title_table_column[0] + '为' + ''.join(table_content1[i * 2]) + "," + title_table_column[1] + '为' + ''.join(table_content1[i * 2 + 1]) + "," +title_table_column[2] + '为' + table_content2[i])
                        elif title_table_column[2] == '操作利率':
                            detail_table_content.insert(len(detail_table_content),'MLF操作情况:' + title_table_column[0] + '为' + ''.join(table_content1[i * 2]) + "," + title_table_column[1] + '为' + ''.join(table_content1[i * 2 + 1]) + "," +title_table_column[2] + '为' + table_content2[i])
                        else:
                            detail_table_content.insert(len(detail_table_content), date + "当日情况存在异常")

                for i in range(0, row):
                    if title_table_column[2] == '中标利率':
                        detail_table_content.insert(len(detail_table_content),
                                                    '逆回购操作情况:' + title_table_column[0] + '为' + ''.join(table_content1[i * 2]) + "," + title_table_column[1] + '为' + ''.join(table_content1[i * 2 + 1]) + "," + title_table_column[2] + '为' + table_content2[i])
                    elif title_table_column[2] == '操作利率':
                        detail_table_content.insert(len(detail_table_content),'MLF操作情况:' + title_table_column[0] + '为' + ''.join(table_content1[i * 2]) + "," + title_table_column[1] + '为' + ''.join(table_content1[i * 2 + 1]) + "," +title_table_column[2] + '为' + table_content2[i])
                    else:
                        detail_table_content.insert(len(detail_table_content), date + "当日情况存在异常")

        if table_alt_1 == []:  # 通知中没有特例表格不做进一步的解析
            pass
        else:
            for t in table_alt_1:
                # 列标题
                pat_title_table_column = '<span style="font-size: medium"><span style="font-family: 宋体"><span style="color: #222222">(.*?)</span>'  # .replace(' ','')
                title_table_column = re.compile(pat_title_table_column).findall(t[1])
                ###表格内容
                # 带汉字
                pat_table_content1 = '<span style="color: #222222">(.*?)</span></span></span><span style="font-family: 宋体"><span style="color: #222222">(.*?)</span>'  # .replace(' ','')
                table_content1 = re.compile(pat_table_content1).findall(t[1])
                # 不带汉字
                pat_table_content2 = '<span style="font-family: arial,sans-serif"><span style="color: #222222"><span style="font-size: medium">(.*?)</span>'  # .replace(' ','')
                table_content2 = re.compile(pat_table_content2).findall(t[1])
                row = len(table_content2)
                for i in range(0, row):
                    if title_table_column[2] == '中标利率':
                        detail_table_content.insert(len(detail_table_content),
                                                    '逆回购操作情况:' + title_table_column[0] + '为' + ''.join(table_content1[i * 2]) + "," + title_table_column[1] + '为' + ''.join(table_content1[i * 2 + 1]) + "," +title_table_column[2] + '为' + table_content2[i])
                    elif title_table_column[2] == '操作利率':
                        detail_table_content.insert(len(detail_table_content),'MLF操作情况:' + title_table_column[0] + '为' + ''.join(table_content1[i * 2]) + "," + title_table_column[1] + '为' + ''.join(table_content1[i * 2 + 1]) + "," + title_table_column[2] + '为' + table_content2[i])
                    else:
                        detail_table_content.insert(len(detail_table_content), date + "当日情况存在异常")

        if table_alt_2 == [] and table_alt_3 == []: # 通知中没有特例表格不做进一步的解析,第二种和第三种特例
            pass
        else:
            for t in table_alt_2 or table_alt_3 :
                # 列标题
                # 默认括号里的内容为group，要匹配括号里的文字要对括号转义
                pat_title_table_column = '<span style="font-size: medium;"><span style="color: rgb\(34, 34, 34\);"><span style="font-family: 宋体;">(.*?)</span></span></span>'  # .replace(' ','')
                title_table_column = re.compile(pat_title_table_column).findall(t[1])
                ###表格内容
                # 带汉字
                pat_table_content1 = '<span style="font-size: medium;"><span style="color: rgb\(34, 34, 34\);"><span style="font-family: arial, sans-serif;">(.*?)</span></span></span><span style="color: rgb\(34, 34, 34\);"><span style="font-family: 宋体;">(.*?)</span></span>'  # .replace(' ','')
                table_content1 = re.compile(pat_table_content1).findall(t[1])
                # 不带汉字
                pat_table_content2 = '<span style="color: rgb\(34, 34, 34\);"><span style="font-family: arial, sans-serif;"><span style="font-size: medium;">(.*?)</span></span></span>'  # .replace(' ','')
                table_content2 = re.compile(pat_table_content2).findall(t[1])
                row = len(table_content2)
                for i in range(0, row):
                    if title_table_column[2] == '中标利率':
                        detail_table_content.insert(len(detail_table_content),'逆回购操作情况:' + title_table_column[0] + '为' + ''.join(table_content1[i * 2]) + "," + title_table_column[1] + '为' + ''.join(table_content1[i * 2 + 1]) + "," +title_table_column[2] + '为' + table_content2[i])
                    elif title_table_column[2] == '操作利率':
                        detail_table_content.insert(len(detail_table_content),'MLF操作情况:' + title_table_column[0] + '为' + ''.join(table_content1[i * 2]) + "," + title_table_column[1] + '为' + ''.join(table_content1[i * 2 + 1]) + "," +title_table_column[2] + '为' + table_content2[i])
                    else:
                        detail_table_content.insert(len(detail_table_content), date + "当日情况存在异常")

        if table_alt_4 == []: # 通知中没有特例表格不做进一步的解析,第4种特例
            pass
        else:
            for t in table_alt_4:
                # 列标题
                # 默认括号里的内容为group，要匹配括号里的文字要对括号转义
                pat_title_table_column = '<span style="color: #222222"><span style="font-family: 宋体">(.*?)</span></span>'  # .replace(' ','')
                title_table_column = re.compile(pat_title_table_column).findall(t[1])
                ###表格内容
                # 带汉字
                pat_table_content1 = '<span style="color: #222222"><span style="font-family: arial, sans-serif">(.*?)</span></span><span style="color: #222222"><span style="font-family: 宋体">(.*?)</span>'  # .replace(' ','')
                table_content1 = re.compile(pat_table_content1).findall(t[1])
                # 不带汉字
                pat_table_content2 = '<span style="color: #222222"><span style="font-family: arial, sans-serif">(.*?)</span>'  # .replace(' ','')
                table_content2 = re.compile(pat_table_content2).findall(t[1])
                row = len(table_content2)
                for i in range(0, row):
                    if title_table_column[2] == '中标利率':
                        detail_table_content.insert(len(detail_table_content),'逆回购操作情况:' + title_table_column[0] + '为' + ''.join(table_content1[i * 2]) + "," + title_table_column[1] + '为' + ''.join(table_content1[i * 2 + 1]) + "," +title_table_column[2] + '为' + table_content2[i])
                    elif title_table_column[2] == '操作利率':
                        detail_table_content.insert(len(detail_table_content),'MLF操作情况:' + title_table_column[0] + '为' + ''.join(table_content1[i * 2]) + "," + title_table_column[1] + '为' + ''.join(table_content1[i * 2 + 1]) + "," +title_table_column[2] + '为' + table_content2[i])
                    else:
                        detail_table_content.insert(len(detail_table_content), date + "当日情况存在异常")



        return date,title_article[0],(detail+';'.join(detail_table_content)).replace('&nbsp;','').replace('<span style="font-size:10.5pt;">','').replace('<span style="color: #222222">','').replace('<span style="font-family: 宋体">','').replace('<span style="font-size: medium">','').replace('</span>','').replace(' ','')

    def toExcel_start(filepath,li_or_tuple):
        wb = openpyxl.load_workbook(filepath)
        worksheet = wb.active
        max = worksheet.max_row
        for x in range(1,len(li_or_tuple)+1):
            worksheet.cell(row=max+1,column=x).value = li_or_tuple[x-1] # 写入新的一行
        wb.save(filepath)
        return print("记录写入成功")

    def toExcel_update(filepath,li_or_tuple):
        wb = openpyxl.load_workbook(filepath)
        worksheet = wb.active
        if worksheet.cell(row=1,column=1).value != li_or_tuple[0]:
            worksheet.insert_rows(2)
            for x in range(1, len(li_or_tuple) + 1):
                worksheet.cell(row=worksheet.max_row + 1, column=x).value = li_or_tuple[x - 1]  # 写入新的一行
            wb.save(filepath)
            return 0
        else:
            return 1

if __name__ == '__main__':
    # 获得公告的总页数
    totalpage = int(re.compile('totalpage="...').findall(Functions.getcode("http://www.pbc.gov.cn/zhengcehuobisi/125207/125213/125431/125475/17081/index1.html"))[0][-3:-1])
    for i in range(45,totalpage+1):
        url = "http://www.pbc.gov.cn/zhengcehuobisi/125207/125213/125431/125475/17081/index"+ str(i) +".html"
        source = Functions.getcode(url)
        pat_link = "/zhengcehuobisi/125207/125213/125431/125475/[0-9]+/index.html"
        linklist = re.compile(pat_link).findall(source)
        for j in range(0,len(linklist)):
            url_detail = "http://www.pbc.gov.cn"+linklist[j]
            source_detail = Functions.getcode(url_detail)
            detail = Functions.getdetail(source_detail)
            print(detail)
            file = r"D:\Python\Project\TEST\spider\forls\OpenMarketOperationRecord.xlsx"
            Functions.toExcel_start(file, detail)
            time.sleep(random.randint(3, 5))

