#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
生成图表
"""

import seaborn as sns
import pandas as pd
import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import openpyxl

# print(plt.rcParams.keys())
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False


path = 'D:\CXM\Paper\利率市场化\图表\图数据_折线.xlsx'

wb = openpyxl.load_workbook(path)
sheet_names = wb.sheetnames
wb.close()




for name in sheet_names:
    df = pd.read_excel(io=path,sheet_name=name,header=0,index_col=0)

    if name == '图表2.7 超额准备金存款利率作为货币市场下限' :
        df.fillna(method="ffill",inplace=True)
        df.plot()
        plt.xticks(rotation=90)
        plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y/%m'))
        plt.gca().xaxis.set_major_locator(mdates.MonthLocator(bymonth=3))

        plt.xlabel('')
        plt.ylabel('')
        # plt.show()
        plt.savefig('D:\CXM\Paper\利率市场化\图表'+'\\'+name+'.png')
    elif name == '图表3.7 银行间与交易所市场回购利率对比':
        df.fillna(method="ffill", inplace=True)
        df.plot()
        plt.xticks(rotation=90)
        plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y/%m'))
        plt.gca().xaxis.set_major_locator(mdates.MonthLocator(bymonth=[11,2,5,8]))

        plt.xlabel('')
        plt.ylabel('')
        # plt.show()
        plt.savefig('D:\CXM\Paper\利率市场化\图表' + '\\' + name + '.png')
    elif name == '图表3.9 贷款利率浮动情况':
        df = pd.read_excel(io=path, sheet_name=name, header=0, converters={'日期':str})
        df.fillna(method="ffill", inplace=True)
        # df.reset_index(inplace=True,col_level='日期')

        print(df)

        wb = openpyxl.load_workbook('D:\\CXM\\123.xlsx')
        ws = wb['Sheet1']
        for i in range(0,df.shape[0]):
            for j in range(0,df.shape[1]):
                ws.cell(row=i+1,column=j+1).value = df.iloc[i, j]
        wb.save('D:\\CXM\\123.xlsx')
        wb.close()
    else:
        df.dropna(inplace=True)
        df.plot()
        plt.xlabel('')
        plt.ylabel('')
        # plt.show()
        plt.savefig('D:\CXM\Paper\利率市场化\图表' + '\\' + name + '.png')
