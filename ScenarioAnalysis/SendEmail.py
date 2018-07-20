# -*- coding: utf-8 -*-
"""
@author: 陈祥明
从VaR日报读取信息并发送邮件
"""
#
import openpyxl
import sys
sys.path.append(r'D:\CXM\Project_New\ScenarioAnalysis')
import Constant
import email.mime.multipart
import email.mime.text
import smtplib

path = Constant.path_var_save1
wb = openpyxl.load_workbook(path)
ws = wb['VaR']
date = Constant.date


msg = email.mime.multipart.MIMEMultipart()
msg_to = ['nizh@dxzq.net.cn', 'wangxj@dxzq.net.cn', 'qiandm@dxzq.net.cn', 'luocx@dxzq.net.cn', 'zhubj@dxzq.net.cn', 'licheng@dxzq.net.cn']
# msg_to = ['314281596@qq.com', '1872917132@qq.com', '570630387@qq.com'] # 测试用
msg['from'] = "chen_xiangming@outlook.com"
msg['to'] = ','.join(msg_to)
msg['subject'] = 'VaR日报'+date


def buildTitle(title, align='center', columns=3):
    return " <tr><td colspan='{0}' style=\"border:1px solid #ccc; text-align:{1}\">{2}</td></tr> ".format(columns, align, title)


def buildCell(value, align='center'):
    return " <td style=\"border: 1px solid #ccc; text-align: {0}\">&nbsp;&nbsp;&nbsp;{1}&nbsp;&nbsp;&nbsp;</td> ".format(align, value)


content = '''
<table style="border-collapse: collapse">
    <tbody>
        ''' + buildTitle('部门整体') + '''
        <tr>
            ''' + buildCell('95%VaR') \
                + buildCell('总市值') \
                + buildCell('敞口') + '''
        </tr>
        <tr>
            ''' + buildCell(str(format(ws.cell(column=1, row=3).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=2, row=3).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=3, row=3).value, '0,.2f')), 'right') + '''
        </tr>
    </tbody>
</table>
<br>
<br/>
<table style="border-collapse: collapse">
    <tbody>
        ''' + buildTitle('组合', columns=4) + '''
        <tr>
            ''' + buildCell('组合名称') \
                + buildCell('95%VaR') \
                + buildCell('总市值') \
                + buildCell('敞口') + '''
        </tr>
        <tr>
            ''' + buildCell(ws.cell(column=1, row=7).value, 'center') \
                + buildCell(str(format(ws.cell(column=2, row=7).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=3, row=7).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=4, row=7).value, '0,.2f')), 'right') + '''
        </tr>
        <tr>
            ''' + buildCell(ws.cell(column=1, row=8).value, 'center') \
                + buildCell(str(format(ws.cell(column=2, row=8).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=3, row=8).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=4, row=8).value, '0,.2f')), 'right') + '''
        </tr>
        <tr>
            ''' + buildCell(ws.cell(column=1, row=9).value, 'center') \
                + buildCell(str(format(ws.cell(column=2, row=9).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=3, row=9).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=4, row=9).value, '0,.2f')), 'right') + '''
        </tr>
        <tr>
            ''' + buildCell(ws.cell(column=1, row=10).value, 'center') \
                + buildCell(str(format(ws.cell(column=2, row=10).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=3, row=10).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=4, row=10).value, '0,.2f')), 'right') + '''
        </tr>
        <tr>
            ''' + buildCell(ws.cell(column=1, row=11).value, 'center') \
                + buildCell(str(format(ws.cell(column=2, row=11).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=3, row=11).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=4, row=11).value, '0,.2f')), 'right') + '''
        </tr>
        <tr>
            ''' + buildCell(ws.cell(column=1, row=12).value, 'center') \
                + buildCell(str(format(ws.cell(column=2, row=12).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=3, row=12).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=4, row=12).value, '0,.2f')), 'right') + '''
        </tr>
        <tr>
            ''' + buildCell(ws.cell(column=1, row=13).value, 'center') \
                + buildCell(str(format(ws.cell(column=2, row=13).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=3, row=13).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=4, row=13).value, '0,.2f')), 'right') + '''
        </tr>
    </tbody>
</table>
<br>
<br/>
<table style="border-collapse: collapse">
    <tbody>
        ''' + buildTitle('期货VaR前三', columns=5) + '''
        <tr>
            ''' + buildCell('品种代码') \
                + buildCell('品种名称') \
                + buildCell('95%VaR') \
                + buildCell('总市值') \
                + buildCell('敞口') + '''
        </tr>
        <tr>
            ''' + buildCell(ws.cell(column=1, row=17).value, 'center') \
                + buildCell(ws.cell(column=2, row=17).value, 'center') \
                + buildCell(str(format(ws.cell(column=3, row=17).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=4, row=17).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=5, row=17).value, '0,.2f')), 'right') + '''
        </tr>
        <tr>
            ''' + buildCell(ws.cell(column=1, row=18).value, 'center') \
                + buildCell(ws.cell(column=2, row=18).value, 'center') \
                + buildCell(str(format(ws.cell(column=3, row=18).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=4, row=18).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=5, row=18).value, '0,.2f')), 'right') + '''
        </tr>
        <tr>
            ''' + buildCell(ws.cell(column=1, row=19).value, 'center') \
                + buildCell(ws.cell(column=2, row=19).value, 'center') \
                + buildCell(str(format(ws.cell(column=3, row=19).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=4, row=19).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=5, row=19).value, '0,.2f')), 'right') + '''
        </tr>
    </tbody>
</table>
<br>
<br/>
<table style="border-collapse: collapse">
    <tbody>
        ''' + buildTitle('股票VaR前三', columns=4) + '''
        <tr>
            ''' + buildCell('股票代码') \
                + buildCell('股票简称') \
                + buildCell('95%VaR') \
                + buildCell('总市值') + '''
        </tr>
        <tr>
            ''' + buildCell(ws.cell(column=1, row=23).value, 'center') \
                + buildCell(ws.cell(column=2, row=23).value, 'center') \
                + buildCell(str(format(ws.cell(column=3, row=23).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=4, row=23).value, '0,.2f')), 'right') + '''
        </tr>
        <tr>
            ''' + buildCell(ws.cell(column=1, row=24).value, 'center') \
                + buildCell(ws.cell(column=2, row=24).value, 'center') \
                + buildCell(str(format(ws.cell(column=3, row=24).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=4, row=24).value, '0,.2f')), 'right') + '''
        </tr>
        <tr>
            ''' + buildCell(ws.cell(column=1, row=25).value, 'center') \
                + buildCell(ws.cell(column=2, row=25).value, 'center') \
                + buildCell(str(format(ws.cell(column=3, row=25).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=4, row=25).value, '0,.2f')), 'right') + '''
        </tr>
    </tbody>
</table>
<br>
<br/>
<table style="border-collapse: collapse">
    <tbody>
        ''' + buildTitle('套保额度使用情况', columns=4) + '''
        <tr>
            ''' + buildCell('', 'center') \
                + buildCell(ws.cell(column=2, row=28).value, 'center') \
                + buildCell(ws.cell(column=3, row=28).value, 'center') \
                + buildCell(ws.cell(column=4, row=28).value, 'center') \
                + '''
        </tr>
        <tr>
            ''' + buildCell(ws.cell(column=1, row=29).value, 'left') \
                + buildCell(str(format(ws.cell(column=2, row=29).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=3, row=29).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=4, row=29).value, '0,.2f')), 'right') \
                + '''
        </tr>
        <tr>
            ''' + buildCell(ws.cell(column=1, row=30).value, 'left') \
                + buildCell(str(format(ws.cell(column=2, row=30).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=3, row=30).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=4, row=30).value, '0,.2f')), 'right') \
                + '''
        </tr>
        <tr>
            ''' + buildCell(ws.cell(column=1, row=31).value, 'left') \
                + buildCell(str(format(ws.cell(column=2, row=31).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=3, row=31).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=4, row=31).value, '0,.2f')), 'right') \
                + '''
        </tr>
        <tr>
            ''' + buildCell(ws.cell(column=1, row=32).value, 'left') \
                + buildCell(str(format(ws.cell(column=2, row=32).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=3, row=32).value, '0,.2f')), 'right') \
                + buildCell(str(format(ws.cell(column=4, row=32).value, '0,.2f')), 'right') \
                + '''
        </tr>
    </tbody>
</table>
<p>
注：<br />
1、 所有VaR根据历史法生成，取过去200个交易日,置信区间为95%<br />
2、  持仓数据来源为当日Nurex的flash持仓<br />
3、 货币基金未纳入统计<br />
</p>
'''




txt = email.mime.text.MIMEText(content, 'html', 'utf-8')
msg.attach(txt)
server = smtplib.SMTP('smtp-mail.outlook.com', 25) # SMTP协议默认端口是25
# server.set_debuglevel(1)

server.ehlo()
server.starttls()


server.login('chen_xiangming@outlook.com', 'cxm032413')
server.sendmail('chen_xiangming@outlook.com', ','.join(msg_to), msg.as_string())
server.quit()

wb.close()