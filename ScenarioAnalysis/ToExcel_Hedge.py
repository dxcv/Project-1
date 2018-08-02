# -*- coding: utf-8 -*-
"""
@author: 陈祥明
对冲
"""
import pandas as pd
import math
import openpyxl
import sys
sys.path.append(r'D:\CXM\Project_New\SQLLINK')
import MSSQL
sys.path.append(r'D:\CXM\Project_New\ScenarioAnalysis')
import Constant
import Calculator

date = Constant.date

dbsa = MSSQL.DB_ScenarioAnalysis()
query = "select * from PositionEod where AccountID in ('77772', '77773', '77774', '77775', '77776', '77777', '77778', '77779') and Date='{}'".format(date)
source = dbsa.ExecQuery(query)

Dataset = []
for i in source:
    a = (i[2], i[3], i[5].encode('latin1').decode('gbk'), i[6], i[7])
    Dataset.append(a)

# position列用于多空轧差
df = pd.DataFrame(Dataset, columns=['InstrumentID', 'Totalposition', 'direction', 'portfolioID', 'parentID'])
df['position'] = df['Totalposition']
df.loc[df['direction'] == '空', 'position'] = -df.loc[df['direction'] == '空', 'Totalposition']

results = df.groupby('InstrumentID').sum()

df_mv = Calculator.PortfolioMvOrExpo(results, date)

SectorID = ('hs300', 'sz50', 'zz500')
SectorInfo = {}
for i in SectorID:
    query = "select InstrumentID from SectorInfo_Stock where SectorID='{}'".format(i)
    SectorInfo[i] = dbsa.ExecQuery(query)

MV_Stock = {}
for m in SectorID:
    df_sum_stock = pd.DataFrame(columns=['Totalposition', 'position', 'MV', 'Expo'])
    for n in SectorInfo[m]:
        if n[0] not in df_mv.index:
            pass
        else:
            df_sum_stock = df_sum_stock.append(df_mv.loc[n[0]])
    df_sum_stock.loc['row_sum'] = df_sum_stock.apply(lambda x: x.sum())
    MV_Stock[m] = df_sum_stock.loc['row_sum']['MV']

Futures = ('IF', 'IC', 'IH')
MV_Future = {'zz500': ['IC'], 'sz50': ['IH'], 'hs300': ['IF']}
for x in SectorID:
    df_sum_future = pd.DataFrame(columns=['Totalposition', 'position', 'MV', 'Expo'])
    for y in df_mv.index:
        if y[:2] == MV_Future[x][0]:
            df_sum_future = df_sum_future.append(df_mv.loc[y])
        else:
            pass

    df_sum_future.loc['row_sum'] = df_sum_future.apply(lambda x: x.sum())
    df_sum_future.fillna(0, inplace=True)
    MV_Future[x].append(-df_sum_future.loc['row_sum']['Expo'])

# print(MV_Future)
# print(MV_Stock)
def hedge(stock, future):
    if stock > future[1]:
        query1 = "select closePrice*multiplier from HistData_Future where tradingDate='{0}' and productID='{1}'".format(date, future[0])
        de = dbsa.ExecQuery(query1)
        query2 = "select Multiplier from Future_Multiplier where RecordDate='{0}' and productID='{1}'".format(date, future[0])
        m = dbsa.ExecQuery(query2)
        num = math.floor((stock - future[1]) / de[0][0] / m[0][0])
        return num
    else:
        return 0

for i in SectorID:
    MV_Future[i].append(hedge(MV_Stock[i], MV_Future[i]))
    print(MV_Future[i])

path = Constant.path_var_save1
wb = openpyxl.load_workbook(path)
ws = wb['VaR']

ws.cell(row=29, column=2).value = MV_Future['zz500'][1]
ws.cell(row=30, column=2).value = MV_Stock['zz500']
ws.cell(row=31, column=2).value = MV_Future['zz500'][2]
if MV_Future['zz500'][2]>0:
    ws.cell(row=32, column=2).value = 0
else:
    ws.cell(row=32, column=2).value = MV_Stock['zz500']-MV_Future['zz500'][1]

ws.cell(row=29, column=3).value = MV_Future['sz50'][1]
ws.cell(row=30, column=3).value = MV_Stock['sz50']
ws.cell(row=31, column=3).value = MV_Future['sz50'][2]
if MV_Future['sz50'][2]:
    ws.cell(row=32, column=3).value = 0
else:
    ws.cell(row=32, column=3).value = MV_Stock['sz50']-MV_Future['sz50'][1]

ws.cell(row=29, column=4).value = MV_Future['hs300'][1]
ws.cell(row=30, column=4).value = MV_Stock['hs300']
ws.cell(row=31, column=4).value = MV_Future['hs300'][2]
if MV_Future['hs300'][2]:
    ws.cell(row=32, column=4).value = 0
else:
    ws.cell(row=32, column=4).value = MV_Stock['hs300'] - MV_Future['hs300'][1]

wb.save(path)
wb.close()
