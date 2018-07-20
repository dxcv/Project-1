# -*- coding: utf-8 -*-
"""
Created on Fri Jul 06 14:03:17 2018

@将结果写入execl

@author: 陈祥明
"""

import openpyxl
import sys
sys.path.append(r'D:\CXM\Project_New\SQLLINK')
import MSSQL
sys.path.append(r'D:\CXM\Project_New\ScenarioAnalysis')
import Constant
# import GetPosition
# import Calculator
# import DataClean

date = Constant.date
dbsa = MSSQL.DB_ScenarioAnalysis()
# ############################################# 将VaR记录输出 #################################
path_VaR = Constant.path_var
path_var_save1 = Constant.path_var_save1
path_var_save2 = Constant.path_var_save2
wb = openpyxl.load_workbook(path_VaR)
ws = wb['VaR']

# 部门整体
query1 = "select VaR_95,Marketvalue,Exposure from VaR_record where portfolioID='ALL' and date='{0}'".format(date)
res1 = dbsa.ExecQuery(query1)
for i in range(0, len(res1[0])):
    ws.cell(row=3, column=i+1).value = res1[0][i]

# 组合
query2 = "select portfolioID,VaR_95,Marketvalue,Exposure from VaR_record where date = '{0}' and productID = 'portfolio' and portfolioID<>'ALL'".format(date)
res2 = dbsa.ExecQuery(query2)
for i in range(0, len(res2)):
    for j in range(0, len(res2[0])):
        ws.cell(row=i+7, column=j+1).value = res2[i][j]

# 期货
query3 = "select top 3 a.productID,b.ProductName, sum(a.VaR_95) as VaR_95,sum(a.Marketvalue) as Marketvalue, sum(a.Exposure) as Exposure from VaR_record a, InstrumentInfo_Future b where a.productID=b.productID and a.date = '{0}' group by a.productID,b.ProductName order by VaR_95".format(date)
res3 = dbsa.ExecQuery(query3)
for i in range(0, len(res3)):
    for j in range(0, len(res3[0])):
        ws.cell(row=i+17, column=j+1).value = res3[i][j]

# 股票
query4 = "select top 3 a.InstrumentID,b.InstrumentName, sum(a.VaR_95) as VaR_95,sum(a.Marketvalue) as Marketvalue from VaR_record a, InstrumentInfo_Stock b where a.InstrumentID=b.InstrumentID and a.date = '{0}' group by a.InstrumentID,b.InstrumentName order by VaR_95".format(date)
res4 = dbsa.ExecQuery(query4)
for i in range(0, len(res4)):
    for j in range(0, len(res4[0])):
        ws.cell(row=i+23, column=j+1).value = res4[i][j]

wb.save(path_var_save1)
wb.save(path_var_save2)
wb.close()


