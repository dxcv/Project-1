# -*- coding: utf-8 -*-
"""
Created on Mon May 14 10:25:54 2018

@author: CXm
"""
#conding = gbk

import openpyxl
import pymssql
import os
import datetime
import time

date = time.strftime('%Y%m%d', time.localtime())
       
conn = pymssql.connect(server='10.63.6.78',user='sa',password='A12345678!',database='OptionContract',charset="utf8")
cursor = conn.cursor()
############################################存量收益互换####################################
cursor.execute("select * from dbo.CustomerStatusReport  where date='{}'".format(date))
wb = openpyxl.load_workbook(u'Z:\\RiskM\\场外风控\\场外业务估值与盯市计算模板_new.xlsx')
results1 = cursor.fetchall()
print ('查询今日存量收益互换成功')
if results1 != []:
    ws1 = wb['存量收益互换']
    for i in range(1,len(results1[0])):
        for j in range(3,len(results1)+3):
            ws1.cell(column=i,row=j).value = results1[j-3][i]
    print ('写入今日存量收益互换成功')
else:
    print ('今日尚无存量收益互换记录')

#############################################已结算收益互换损益#####################################
#重置查询
cursor.execute('select * from dbo.CustomerDeposit')
results2 = cursor.fetchall()
print ('查询保证金流水记录成功')
ws2 = wb['已结算收益互换损益']
for i in range(1,len(results2[0])-2):
    for j in range(3,len(results2)+3):
        ws2.cell(column=i,row=j).value = results2[j-3][i]
        ws2.cell(column=7,row=j).value = -results2[j-3][6]
print ('写入保证金流水记录成功')
conn.close()
wb.save('Z:\\RiskM\\场外风控\\场外业务估值与盯市计算模板_new.xlsx')
wb.close()
print ('数据已更新至场外业务估值与盯市计算模板_new')


##另存到整体数据文件夹
if results1==[]:
    print ('今日存量收益互换记录尚未更新，无法生成日报表')
else:
    wb_other = openpyxl.load_workbook(u'Z:\\RiskM\\场外风控\\场外业务估值与盯市计算模板_new.xlsx')

    path_list = [u'Z:\\RiskM\\场外风控\\整体数据\\',date,u'衍生品部上海团队场外业务风控指标日报表.xlsx']
    path = ''.join(path_list)
    wb_other.save(path)
    wb_other.close()
    print ('日报表已保存至整体数据文件夹')
    
###暂停查看结果    
# os.system("pause")
